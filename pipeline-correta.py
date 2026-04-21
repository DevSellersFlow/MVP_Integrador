"""
pipeline.py
===========
Orquestrador principal do SellersFlow.

Une Reader → Mapper → (AIEngine) → Filler em um pipeline único.
É o único ponto de entrada que o app.py (Streamlit) precisa chamar.

Design:
  - Stateless: cada chamada a run() é independente
  - Retorna PipelineResult com todos os artefatos e logs
  - Suporta modo "dry_run" (mapeia mas não grava arquivo)
  - Suporta enriquecimento por IA (opt-in)
"""

from __future__ import annotations

import io
import logging
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd

from core.reader import AmazonSheetReader, AmazonReadResult
from core.mapper import ColumnMapper, MappingResult
from core.filler import MarketplaceFiller, FillResult
from ai.ai_engine import AIEngine

logger = logging.getLogger(__name__)

# ─── Caminho do banco de aprendizado ─────────────────────────────────────────

DEFAULT_DB_PATH = Path(__file__).parent / "data" / "mappings_db" / "learned.json"


# ─── Dataclass de resultado ───────────────────────────────────────────────────

@dataclass
class PipelineResult:
    # Metadados
    marketplace: str
    elapsed_seconds: float

    # Resultados por etapa
    read_result: Optional[AmazonReadResult] = None
    mapping_result: Optional[MappingResult] = None
    fill_result: Optional[FillResult] = None

    # Flags de alto nível
    success: bool = False
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def output_path(self) -> Optional[str]:
        return self.fill_result.output_path if self.fill_result else None

    @property
    def amazon_df(self) -> Optional[pd.DataFrame]:
        return self.read_result.df if self.read_result else None

    @property
    def has_errors(self) -> bool:
        return bool(self.errors)


# ─── Pipeline ─────────────────────────────────────────────────────────────────

class SellersFlowPipeline:
    """
    Pipeline principal do SellersFlow.

    Uso típico:
        pipeline = SellersFlowPipeline()
        result = pipeline.run(
            amazon_file=bytes_io_amazon,
            template_file=bytes_io_template,
            marketplace="Shopee",
            use_ai=True,
        )
        if result.success:
            print(result.output_path)
    """

    def __init__(
        self,
        db_path: Optional[Path] = None,
        output_dir: Optional[str] = None,
    ):
        self._reader = AmazonSheetReader()
        self._mapper = ColumnMapper(db_path=db_path or DEFAULT_DB_PATH)
        self._filler = MarketplaceFiller()
        self._ai = AIEngine()
        self._output_dir = output_dir

    # ── Pública ───────────────────────────────────────────────────────────────

    def run(
        self,
        amazon_file,
        template_file,
        marketplace: str,
        use_ai: bool = False,
        enrich_ai: bool = False,
        dry_run: bool = False,
    ) -> PipelineResult:
        """
        Executa o pipeline completo.

        Args:
            amazon_file: BytesIO da planilha Amazon.
            template_file: BytesIO do template do marketplace.
            marketplace: "Shopee" | "Temu" | "Vendor".
            use_ai: Se True, usa IA como fallback de mapeamento.
            enrich_ai: Se True, aplica enriquecimento de conteúdo via IA.
            dry_run: Se True, não grava arquivo de saída.

        Returns:
            PipelineResult completo.
        """
        t0 = time.perf_counter()
        result = PipelineResult(marketplace=marketplace, elapsed_seconds=0.0)

        # ── Etapa 1: Leitura Amazon ───────────────────────────────────────
        logger.info("[Pipeline] Lendo Amazon...")
        read_result = self._reader.read(amazon_file)
        result.read_result = read_result

        if read_result.has_errors:
            result.errors.extend(read_result.errors)
            result.elapsed_seconds = time.perf_counter() - t0
            return result

        result.warnings.extend(read_result.warnings)
        amazon_df = read_result.df

        # ── Etapa 2: Enriquecimento por IA (opcional) ─────────────────────
        if enrich_ai:
            logger.info("[Pipeline] Enriquecendo com IA...")
            amazon_df = self._apply_enrichment(amazon_df, marketplace)

        # ── Etapa 3: Obter cabeçalhos do template ─────────────────────────
        _template_ext = None
        if hasattr(template_file, "name") and template_file.name:
            from pathlib import Path as _P
            _template_ext = _P(template_file.name).suffix.lower() or None
        if hasattr(template_file, "seek"):
            template_file.seek(0)
            _template_bytes = template_file.read()
        else:
            _template_bytes = Path(template_file).read_bytes()

        logger.info("[Pipeline] Lendo template %s...", marketplace)
        dest_headers = self._read_template_headers(_template_bytes, marketplace)
        if dest_headers is None:
            result.errors.append("Não foi possível ler os cabeçalhos do template.")
            result.elapsed_seconds = time.perf_counter() - t0
            return result

        # ── Etapa 4: Mapeamento ───────────────────────────────────────────
        logger.info("[Pipeline] Construindo mapeamento...")
        ai_engine = self._ai if use_ai else None
        mapping = self._mapper.build_mapping(
            amazon_df=amazon_df,
            dest_headers=dest_headers,
            marketplace=marketplace,
            ai_engine=ai_engine,
        )
        result.mapping_result = mapping

        if dry_run:
            result.success = True
            result.elapsed_seconds = time.perf_counter() - t0
            return result

        # ── Etapa 5: Preenchimento ────────────────────────────────────────
        logger.info("[Pipeline] Preenchendo template...")
        fill_result = self._filler.fill(
            amazon_df=amazon_df,
            mapping=mapping,
            template_file=io.BytesIO(_template_bytes),
            output_dir=self._output_dir,
            template_ext=_template_ext,
        )
        result.fill_result = fill_result

        if fill_result.has_errors:
            result.errors.extend(fill_result.errors)
        else:
            result.success = True

        result.warnings.extend(fill_result.warnings)
        result.elapsed_seconds = round(time.perf_counter() - t0, 2)
        logger.info("[Pipeline] Concluído em %.2fs", result.elapsed_seconds)
        return result

    def learn_mapping(
        self, marketplace: str, dest_col: str, source_col: str
    ) -> None:
        """Persiste uma decisão de mapeamento confirmada pelo usuário."""
        self._mapper.learn(marketplace, dest_col, source_col)

    # ── Privadas ──────────────────────────────────────────────────────────────

    def _read_template_headers(
        self, template_bytes: bytes, marketplace: str
    ) -> Optional[dict[int, str]]:
        import os
        import tempfile
        from openpyxl import load_workbook

        # Config inline garante funcionamento mesmo se filler.py não estiver atualizado
        _INLINE_CONFIG: dict[str, dict] = {
            "Mercado Livre": {"sheet_index_after_ajuda": True, "header_row": 3},
        }

        try:
            from core.filler import MARKETPLACE_CONFIG
            config = MARKETPLACE_CONFIG.get(marketplace) or _INLINE_CONFIG.get(marketplace)
        except Exception:
            config = _INLINE_CONFIG.get(marketplace)

        if not config:
            logger.error(
                "Marketplace '%s' não encontrado em MARKETPLACE_CONFIG. "
                "Verifique se core/filler.py está atualizado com este marketplace.",
                marketplace,
            )
            return None

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(template_bytes)
                tmp_path = tmp.name

            wb = load_workbook(tmp_path, read_only=True)
            logger.info("Template '%s' — abas disponíveis: %s", marketplace, wb.sheetnames)

            if marketplace == "Vendor":
                prefix = config.get("sheet_prefix", "Modelo-")
                ws = next(
                    (wb[s] for s in wb.sheetnames if s.startswith(prefix)),
                    wb.active,
                )
            elif config.get("sheet_index_after_ajuda"):
                # Mercado Livre: terceira aba (índice 2) — nome varia por categoria
                if len(wb.sheetnames) >= 3:
                    ws = wb[wb.sheetnames[2]]
                    logger.info("Mercado Livre: usando aba '%s' (índice 2)", ws.title)
                else:
                    ws = wb[wb.sheetnames[-1]]
                    logger.warning("Mercado Livre: menos de 3 abas; usando '%s'", ws.title)
            else:
                sheet = config.get("sheet", "")
                ws = wb[sheet] if sheet in wb.sheetnames else wb.active

            headers = {}
            for cell in ws[config["header_row"]]:
                if cell.value:
                    headers[cell.column] = cell.value

            logger.info(
                "Cabeçalhos lidos: %d colunas na aba '%s'", len(headers), ws.title
            )
            wb.close()
            return headers

        except Exception as exc:
            logger.error(
                "Erro ao ler cabeçalhos do template '%s': %s",
                marketplace, exc, exc_info=True,
            )
            return None
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    def _apply_enrichment(
        self, df: pd.DataFrame, marketplace: str
    ) -> pd.DataFrame:
        """Aplica enriquecimento de IA linha por linha (com cache automático)."""
        enriched_rows = []
        for _, row in df.iterrows():
            row_dict = row.dropna().to_dict()
            enrich = self._ai.enrich_row(row_dict, marketplace)
            if enrich:
                # Sobrescreve campos enriquecidos
                for col_alias, field_key in [
                    ("item name", "title"),
                    ("nome do produto", "title"),
                    ("product description", "description"),
                    ("descrição do produto", "description"),
                ]:
                    if col_alias in row.index and field_key in enrich:
                        row[col_alias] = enrich[field_key]
                # Bullet points → colunas existentes
                bullets = enrich.get("bullets", [])
                for i, bullet in enumerate(bullets[:5], start=1):
                    col_bp = f"bullet point{i}" if i > 1 else "bullet point"
                    if col_bp in row.index:
                        row[col_bp] = bullet
            enriched_rows.append(row)

        return pd.DataFrame(enriched_rows)